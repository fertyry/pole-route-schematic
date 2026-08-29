from __future__ import annotations

from dataclasses import replace

from openpyxl import Workbook

from pole_route.domain.pea_asset import (
    AssetMatchState,
    AssetSideRelation,
    PEAAsset,
    PEAAssetType,
    merge_pea_assets,
)
from pole_route.domain.pea_gis import PEAPoleRecord
from pole_route.domain.pea_ordering import PEAPoleOrdering, PEAPoleReviewEntry, PoleQCStatus
from pole_route.domain.route import GeoPoint, Route
from pole_route.geometry.pea_asset_matching import match_pea_assets
from pole_route.importers.pea_assets import import_pea_assets
from pole_route.importers.pea_gis import discover_pea_workbook, import_ds_poles
from pole_route.project.storage import (
    load_project_file,
    pea_asset_matches_from_data,
    pea_asset_matches_to_data,
    pea_assets_from_data,
    pea_assets_to_data,
    save_project_file,
)


def _workbook(path):
    workbook = Workbook()
    transformer = workbook.active
    transformer.title = "DS_Transformer"
    transformer.append(["Transformer ID", "Latitude", "Longitude", "Capacity", "Phase", "Custom"])
    transformer.append(["TX-1", 13.0, 100.0, "250 kVA", "3", "เก็บไว้"])
    transformer.append([None, "bad", None, None, None, "audit"])
    switch = workbook.create_sheet("DS_Switch")
    switch.append(["Switch ID", "Lat", "Lng", "Switch Type", "Status"])
    switch.append(["SW-1", 13.0001, 100.0001, "LBS", "Active"])
    workbook.create_sheet("DS_Capacitor")
    workbook.save(path)


def _pole(source_id, longitude, *, row=2):
    return PEAPoleRecord(source_id, "DS_Pole", row, 13.0, longitude)


def _asset(stable_id="DS_Transformer:tx-1", longitude=100.0):
    return PEAAsset(stable_id, "DS_Transformer", 2, PEAAssetType.TRANSFORMER,
                    "TX-1", 13.0, longitude, {"Custom": "เก็บไว้"})


def _route(reverse=False):
    points = (GeoPoint(100.0, 13.0), GeoPoint(100.01, 13.0))
    return Route("Main", "test", tuple(reversed(points)) if reverse else points)


def test_asset_profiles_discovery_parse_qc_and_raw_attributes(tmp_path):
    path = tmp_path / "pea.xlsx"
    _workbook(path)
    discovery = discover_pea_workbook(path)
    assert [item.profile for item in discovery.supported_sheets] == ["DS_Transformer", "DS_Switch"]
    assert discovery.unsupported_ds_sheets == ("DS_Capacitor",)
    assets = import_pea_assets(path)
    assert len(assets) == 3
    transformer = assets[0]
    assert transformer.stable_id == "DS_Transformer:tx-1"
    assert transformer.rating == "250 kVA" and transformer.phase == "3"
    assert transformer.raw_attributes["Custom"] == "เก็บไว้"
    invalid = assets[1]
    assert not invalid.coordinate_valid and len(invalid.qc_warnings) == 3
    assert "fingerprint:" in invalid.stable_id
    switch = assets[2]
    assert switch.asset_type is PEAAssetType.SWITCH
    assert switch.equipment_subtype == "LBS" and switch.status == "Active"


def test_real_schema_aliases_are_source_neutral_and_preserve_coordinates(tmp_path):
    path = tmp_path / "real-schema.xlsx"
    workbook = Workbook()
    pole = workbook.active
    pole.title = "DS_Pole"
    pole.append(["รหัส TAG", "LATITUDE", "LONGITUDE", "ความสูงเสา", "ระดับแรงดัน"])
    pole.append(["22PL-1", 16.7, 103.3, "12 เมตร", "22 kV"])
    transformer = workbook.create_sheet("DS_Transformer")
    transformer.append([
        "PEANO หม้อแปลง", "LATITUDE", "LONGITUDE", "ระดับแรงดัน",
        "ค่าพิกัด kVA หม้อแปลง", "เฟสที่ติดตั้ง", "ประเภทการติดตั้ง",
        "สถานะการก่อสร้าง", "รหัสสายป้อนที่ 1",
    ])
    transformer.append([
        "68-001", 16.700001234567, 103.300001234567, "22 kV", 100,
        "ABC", "บนเสา", "Installed/Existing", "KRA01",
    ])
    switch = workbook.create_sheet("DS_Switch")
    switch.append([
        "รหัสอุปกรณ์", "LATITUDE", "LONGITUDE", "ระดับแรงดัน",
        "ประเภทย่อยของสวิตช์", "เฟสที่ติดตั้ง", "สถานะปัจจุบัน",
    ])
    switch.append(["KRA-SW-1", 16.71, 103.31, "22 kV", "Fuse Dropout", "ABC", "Close"])
    workbook.save(path)

    poles = import_ds_poles(path)
    assets = import_pea_assets(path)

    assert poles[0].source_id == "22PL-1"
    tx, switch_asset = assets
    assert tx.source_asset_id == "68-001"
    assert (tx.latitude, tx.longitude) == (16.700001234567, 103.300001234567)
    assert tx.rating == "100" and tx.phase == "ABC"
    assert tx.equipment_subtype == "บนเสา" and tx.feeder_reference == "KRA01"
    assert switch_asset.source_asset_id == "KRA-SW-1"
    assert switch_asset.equipment_subtype == "Fuse Dropout"


def test_asset_identity_survives_row_order_change(tmp_path):
    first = tmp_path / "first.xlsx"; second = tmp_path / "second.xlsx"
    _workbook(first); _workbook(second)
    workbook = Workbook(); sheet = workbook.active; sheet.title = "DS_Transformer"
    sheet.append(["Transformer ID", "Latitude", "Longitude"])
    sheet.append(["OTHER", 13.1, 100.1]); sheet.append(["TX-1", 13.0, 100.0]); workbook.save(second)
    one = import_pea_assets(first)[0]
    two = next(item for item in import_pea_assets(second) if item.source_asset_id == "TX-1")
    assert one.stable_id == two.stable_id and one.source_row != two.source_row


def test_matcher_suggests_deterministically_and_keeps_suggestion_unconfirmed():
    asset = _asset(longitude=100.00001)
    poles = [_pole("B", 100.00002, row=3), _pole("A", 100.00002, row=2)]
    match = match_pea_assets([asset], poles)[0]
    assert match.state is AssetMatchState.AMBIGUOUS
    assert [item.pole_id for item in match.candidates[:2]] == ["A", "B"]
    assert match.suggested_pole_key is not None and match.confirmed_pole_key is None
    assert match.candidates[0].strength == "strong"


def test_matcher_invalid_no_poles_and_excluded_pole_handling():
    invalid = replace(_asset(), latitude=None)
    assert match_pea_assets([invalid], [ _pole("A", 100.0) ])[0].state is AssetMatchState.UNMATCHED
    pole = _pole("A", 100.0)
    entry = PEAPoleReviewEntry(pole.source_key, "A", 0, 0, 13, 100, 1, 1, None, False, PoleQCStatus.NORMAL)
    ordering = PEAPoleOrdering((entry,))
    match = match_pea_assets([_asset()], [pole], ordering)[0]
    assert match.candidates[0].pole_included is False
    assert match.state is AssetMatchState.UNMATCHED
    assert match_pea_assets([_asset()], [])[0].state is AssetMatchState.UNMATCHED


def test_side_evidence_is_supporting_only_and_route_reversal_invariant():
    asset = replace(_asset(), latitude=13.00005, longitude=100.005)
    same_pole = replace(_pole("SAME", 100.005), latitude=13.00006)
    opposite_pole = replace(_pole("OPPOSITE", 100.005, row=3), latitude=12.99998)

    forward = match_pea_assets(
        [asset], [same_pole, opposite_pole], main_route=_route()
    )[0]
    reverse = match_pea_assets(
        [asset], [same_pole, opposite_pole], main_route=_route(True)
    )[0]

    by_id = {candidate.pole_id: candidate for candidate in forward.candidates}
    reversed_by_id = {candidate.pole_id: candidate for candidate in reverse.candidates}
    assert by_id["SAME"].side_relation is AssetSideRelation.SAME_SIDE
    assert by_id["OPPOSITE"].side_relation is AssetSideRelation.OPPOSITE_SIDE
    assert reversed_by_id["SAME"].side_relation is AssetSideRelation.SAME_SIDE
    assert reversed_by_id["OPPOSITE"].side_relation is AssetSideRelation.OPPOSITE_SIDE
    assert forward.state is not AssetMatchState.CONFIRMED


def test_side_evidence_uses_centerline_dead_band():
    asset = replace(_asset(), latitude=13.000001, longitude=100.005)
    pole = replace(_pole("A", 100.005), latitude=13.00005)
    match = match_pea_assets([asset], [pole], main_route=_route())[0]
    assert match.candidates[0].side_relation is AssetSideRelation.UNCERTAIN


def test_manual_confirmation_survives_recompute_and_reimport():
    poles = [_pole("A", 100.0, row=2), _pole("B", 100.0001, row=3)]
    suggested = match_pea_assets([_asset()], poles)[0]
    confirmed = suggested.confirm(poles[1].source_key)
    recomputed = match_pea_assets([_asset()], poles, previous=[confirmed])[0]
    assert recomputed.state is AssetMatchState.CONFIRMED
    assert recomputed.confirmed_pole_key == poles[1].source_key and recomputed.manual_override
    refreshed = replace(_asset(), rating="500 kVA", source_row=9)
    merged = merge_pea_assets([_asset()], [confirmed], [refreshed], imported_sheets={"DS_Transformer"})
    assert len(merged.assets) == 1 and merged.assets[0].rating == "500 kVA"
    assert merged.matches[0] == confirmed


def test_reimport_marks_missing_and_adds_new_without_duplicate():
    existing = _asset()
    new = replace(_asset("DS_Transformer:tx-2"), source_asset_id="TX-2")
    merged = merge_pea_assets([existing], [], [new], imported_sheets={"DS_Transformer"})
    assert (merged.added, merged.updated, merged.missing_from_source) == (1, 0, 1)
    assert len(merged.assets) == 2
    assert next(item for item in merged.assets if item.stable_id == existing.stable_id).source_present is False


def test_asset_and_match_project_round_trip_and_legacy_default(tmp_path):
    asset = replace(_asset(), latitude=13.00005)
    pole = replace(_pole("A", 100.0), latitude=13.00006)
    match = match_pea_assets([asset], [pole], main_route=_route())[0].confirm(
        pole.source_key
    )
    assert match.candidates[0].side_relation is AssetSideRelation.SAME_SIDE
    path = tmp_path / "new.prs"
    save_project_file(path, {"pea_assets": pea_assets_to_data([asset]),
                             "pea_asset_matches": pea_asset_matches_to_data([match])})
    document = load_project_file(path)
    assert pea_assets_from_data(document["pea_assets"]) == [asset]
    assert pea_asset_matches_from_data(document["pea_asset_matches"]) == [match]
    legacy = tmp_path / "old.prs"; save_project_file(legacy, {})
    old = load_project_file(legacy)
    assert old["pea_assets"] == [] and old["pea_asset_matches"] == []
