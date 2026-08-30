from openpyxl import Workbook
from PySide6.QtWidgets import QDialog, QToolBar

from pole_route.importers.asset_importer import (
    HEADER_ALIASES as ASSET_ALIASES,
)
from pole_route.importers.asset_importer import (
    REQUIRED_FIELDS as ASSET_REQUIRED,
)
from pole_route.importers.asset_importer import (
    assets_from_table,
    inspect_asset_file,
    suggest_asset_mapping,
)
from pole_route.importers.pea_assets import (
    DS_SWITCH_PROFILE,
    DS_TRANSFORMER_PROFILE,
    import_pea_assets,
)
from pole_route.importers.pea_gis import discover_pea_workbook, import_ds_poles
from pole_route.importers.pole_importer import (
    HEADER_ALIASES as POLE_ALIASES,
)
from pole_route.importers.pole_importer import (
    REQUIRED_FIELDS as POLE_REQUIRED,
)
from pole_route.importers.pole_importer import (
    inspect_pole_file,
    poles_from_table,
    suggest_column_mapping,
)
from pole_route.importers.tabular_source import detect_header, list_worksheets
from pole_route.ui.main_window import MainWindow
from pole_route.ui.pea_sheet_selection_dialog import PEASheetSelectionDialog
from pole_route.ui.tabular_source_dialog import TabularSourceDialog


def _multisheet_workbook(path):
    workbook = Workbook()
    cover = workbook.active
    cover.title = "Cover"
    cover.append(["Project report"])
    pole = workbook.create_sheet("Poles")
    pole.append(["รายงานเสาไฟฟ้า"])
    pole.append(["Project", "Demo"])
    pole.append(["Date", "2026-08-30"])
    pole.append([])
    pole.append(["Pole ID", "LAT", "LNG", "Description", "Side", "Quantity"])
    pole.append(["P-1", 13.1, 100.1, "ต้นทดสอบ", "R", 2])
    asset = workbook.create_sheet("Assets")
    asset.append(["Asset report"])
    asset.append([])
    asset.append(["Asset ID", "Type", "Latitude", "Longitude", "Description"])
    asset.append(["TX-1", "Transformer", 13.1, 100.1, "หม้อแปลง"])
    workbook.save(path)


def test_multisheet_pole_selects_second_sheet_and_detects_row_five(tmp_path):
    path = tmp_path / "multi.xlsx"
    _multisheet_workbook(path)
    assert list_worksheets(path) == ("Cover", "Poles", "Assets")
    table = inspect_pole_file(path, sheet_name="Poles")
    assert table.header_row == 5 and table.confidence == "High"
    poles = poles_from_table(table, suggest_column_mapping(table.headers))
    assert poles[0].number == "P-1" and poles[0].side.value == "Right"
    assert poles[0].installed_quantity == 2


def test_multisheet_asset_selects_third_sheet_and_detects_header(tmp_path):
    path = tmp_path / "multi.xlsx"
    _multisheet_workbook(path)
    table = inspect_asset_file(path, sheet_name="Assets")
    assert table.header_row == 3 and table.confidence == "High"
    assets = assets_from_table(table, suggest_asset_mapping(table.headers))
    assert assets[0].source_asset_id == "TX-1"


def test_manual_header_override_changes_preview_table(tmp_path):
    path = tmp_path / "manual.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Pole ID", "Latitude", "Longitude"])
    sheet.append(["Example", "Latitude", "Longitude"])
    sheet.append(["P-1", 13.0, 100.0])
    workbook.save(path)
    automatic = inspect_pole_file(path)
    manual = inspect_pole_file(path, header_row=2)
    assert automatic.header_row == 1
    assert manual.header_row == 2 and manual.confidence == "Manual"
    assert manual.rows[0][0] == "P-1"


def test_header_detection_prefers_complete_coordinate_table_after_metadata():
    rows = [
        ("Pole ID", "Notes"),
        ("Project", "Demo"),
        (),
        ("ลำดับ", "พิกัดละติจูด", "พิกัดลองจิจูด", "รายละเอียด"),
        (1, 13.0, 100.0, "เสา"),
    ]
    aliases = {**POLE_ALIASES}
    aliases["latitude"] = {*aliases["latitude"], "พิกัดละติจูด"}
    aliases["longitude"] = {*aliases["longitude"], "พิกัดลองจิจูด"}
    detection = detect_header(rows, aliases, POLE_REQUIRED)
    assert detection.row_number == 4 and detection.confidence == "High"


def test_low_confidence_when_required_fields_are_missing():
    detection = detect_header(
        [("Report", "Date"), ("Asset ID", "Description")],
        ASSET_ALIASES,
        ASSET_REQUIRED,
    )
    assert detection.row_number == 2 and detection.confidence == "Low"


def test_xlsm_is_listed_and_read_without_running_macros(tmp_path):
    path = tmp_path / "data.xlsm"
    workbook = Workbook()
    workbook.active.title = "Data"
    workbook.active.append(["Pole ID", "Latitude", "Longitude"])
    workbook.active.append(["P-1", 13.0, 100.0])
    workbook.save(path)
    assert list_worksheets(path) == ("Data",)
    assert inspect_pole_file(path, sheet_name="Data").rows[0][0] == "P-1"


def test_source_dialog_allows_sheet_and_header_override(qtbot, tmp_path):
    path = tmp_path / "multi.xlsx"
    _multisheet_workbook(path)
    dialog = TabularSourceDialog(path, POLE_ALIASES, POLE_REQUIRED, {
        "number": "Pole No.", "latitude": "Latitude", "longitude": "Longitude"
    })
    qtbot.addWidget(dialog)
    dialog.sheet_selector.setCurrentIndex(dialog.sheet_selector.findData("Poles"))
    assert dialog.header_row_number() == 5
    dialog.header_row.setValue(5)
    assert "High" in dialog.confidence.text()
    assert "Pole No." in dialog.mapping_summary.text()


def test_source_dialog_allows_unrecognized_header_for_manual_mapping(qtbot, tmp_path):
    path = tmp_path / "custom_headers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Custom A", "Custom B", "Custom C"])
    sheet.append(["P-1", 13.0, 100.0])
    workbook.save(path)

    dialog = TabularSourceDialog(path, POLE_ALIASES, POLE_REQUIRED, {
        "number": "Pole No.", "latitude": "Latitude", "longitude": "Longitude"
    })
    qtbot.addWidget(dialog)
    dialog._accept_selection()

    assert dialog.result() == QDialog.DialogCode.Accepted


def _pea_workbook(path):
    workbook = Workbook()
    workbook.active.title = "DS_Pole"
    workbook.active.append(["รหัส TAG", "LATITUDE", "LONGITUDE", "ความสูง", "แรงดัน"])
    workbook.active.append(["P-1", 13.0, 100.0, 12, "22 kV"])
    transformer = workbook.create_sheet("DS_Transformer")
    transformer.append(["PEANO หม้อแปลง", "LATITUDE", "LONGITUDE"])
    transformer.append(["TX-1", 13.0, 100.0])
    switch = workbook.create_sheet("DS_Switch")
    switch.append(["รหัสอุปกรณ์", "LATITUDE", "LONGITUDE"])
    switch.append(["SW-1", 13.0, 100.0])
    workbook.create_sheet("DS_MVConductor")
    workbook.create_sheet("Notes")
    workbook.save(path)


def test_pea_sheet_selection_and_subset_import(qtbot, tmp_path):
    path = tmp_path / "pea.xlsx"
    _pea_workbook(path)
    discovery = discover_pea_workbook(path)
    dialog = PEASheetSelectionDialog(discovery)
    qtbot.addWidget(dialog)
    assert {"DS_Pole", "DS_Transformer", "DS_Switch"} <= dialog.selected_sheet_names()
    assert not dialog._checks["DS_MVConductor"].isEnabled()
    dialog._checks["DS_Pole"].setChecked(False)
    dialog._checks["DS_Switch"].setChecked(False)
    assert dialog.selected_sheet_names() == {"DS_Transformer"}
    assert len(import_pea_assets(path, (DS_TRANSFORMER_PROFILE,))) == 1
    assert len(import_pea_assets(path, (DS_SWITCH_PROFILE,))) == 1
    assert len(import_ds_poles(path, "DS_Pole")) == 1


def test_primary_workflow_actions_are_source_neutral_and_wired(qtbot):
    window = MainWindow()
    qtbot.addWidget(window)
    assert window.review_pea_order_action.text() == "Review pole order"
    assert window.check_google_earth_action.text() == "Check Pole QC"
    assert window.check_pea_assets_google_earth_action.text() == "Check Asset QC"
    toolbar = window.findChild(QToolBar, "importQcToolbar")
    assert toolbar is not None
    actions = {
        window.import_poles_action,
        window.import_assets_action,
        window.import_pea_gis_action,
        window.review_pea_order_action,
        window.review_pea_assets_action,
        window.check_google_earth_action,
        window.check_pea_assets_google_earth_action,
    }
    assert actions <= set(toolbar.actions())
