"""Inspect, map, import, and validate pole data from CSV and Excel."""

import csv
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from pole_route.domain.pole import Pole, PoleSide

FIELD_LABELS = {
    "number": "Pole No.",
    "latitude": "Latitude",
    "longitude": "Longitude",
    "detail": "Detail",
    "side": "Side",
}
REQUIRED_FIELDS = ("number", "latitude", "longitude")
OPTIONAL_FIELDS = ("detail", "side")
SUPPORTED_SUFFIXES = {".csv", ".xlsx"}
HEADER_SCAN_LIMIT = 20

HEADER_ALIASES = {
    "number": {
        "poleno",
        "polenumber",
        "poleid",
        "pole",
        "no",
        "number",
        "หมายเลขเสา",
        "เลขเสา",
    },
    "latitude": {"latitude", "lat", "ละติจูด"},
    "longitude": {"longitude", "long", "lon", "lng", "ลองจิจูด"},
    "detail": {"detail", "details", "description", "remark", "remarks", "รายละเอียด"},
    "side": {"side", "roadside", "ฝั่ง", "ด้าน"},
}


class PoleImportError(ValueError):
    """A source file cannot be converted into valid pole records."""


@dataclass(frozen=True, slots=True)
class PoleTable:
    """Raw tabular content selected from a source file."""

    headers: tuple[str, ...]
    rows: tuple[tuple[object, ...], ...]
    header_row: int


def inspect_pole_file(path: str | Path) -> PoleTable:
    """Read a supported source and detect the most likely header row."""
    source = Path(path)
    suffix = source.suffix.casefold()
    if suffix not in SUPPORTED_SUFFIXES:
        raise PoleImportError("Choose a .csv or .xlsx pole-data file")
    if not source.is_file():
        raise PoleImportError(f"File not found: {source}")

    raw_rows = _read_csv_rows(source) if suffix == ".csv" else _read_xlsx_rows(source)
    if not raw_rows:
        raise PoleImportError("The file is empty")
    header_index = _detect_header_index(raw_rows)
    headers = _unique_headers(raw_rows[header_index])
    data_rows = tuple(
        tuple(row[index] if index < len(row) else None for index in range(len(headers)))
        for row in raw_rows[header_index + 1 :]
        if any(value not in (None, "") for value in row)
    )
    return PoleTable(headers, data_rows, header_index + 1)


def suggest_column_mapping(headers: tuple[str, ...]) -> dict[str, str | None]:
    """Match common English and Thai header variants to application fields."""
    mapping: dict[str, str | None] = {}
    for field in (*REQUIRED_FIELDS, *OPTIONAL_FIELDS):
        matches = [header for header in headers if _normalize_header(header) in HEADER_ALIASES[field]]
        mapping[field] = matches[0] if len(matches) == 1 else None
    return mapping


def import_poles(
    path: str | Path,
    mapping: dict[str, str | None] | None = None,
) -> list[Pole]:
    """Import poles using automatic aliases or an explicit column mapping."""
    table = inspect_pole_file(path)
    selected_mapping = mapping or suggest_column_mapping(table.headers)
    return poles_from_table(table, selected_mapping)


def poles_from_table(table: PoleTable, mapping: dict[str, str | None]) -> list[Pole]:
    """Convert inspected rows using a user-confirmed column mapping."""
    missing = [FIELD_LABELS[field] for field in REQUIRED_FIELDS if not mapping.get(field)]
    if missing:
        raise PoleImportError("Choose columns for: " + ", ".join(missing))

    header_indexes = {header: index for index, header in enumerate(table.headers)}
    unknown = [column for column in mapping.values() if column and column not in header_indexes]
    if unknown:
        raise PoleImportError("Mapped columns not found: " + ", ".join(unknown))

    poles = [
        _row_to_pole(row, table.header_row + row_offset, mapping, header_indexes)
        for row_offset, row in enumerate(table.rows, start=1)
    ]
    if not poles:
        raise PoleImportError("The file contains headers but no pole rows")
    return poles


def _read_csv_rows(path: Path) -> list[tuple[object, ...]]:
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [tuple(row) for row in csv.reader(handle)]
    except UnicodeDecodeError as error:
        raise PoleImportError("CSV must be saved with UTF-8 encoding") from error


def _read_xlsx_rows(path: Path) -> list[tuple[object, ...]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as error:
        raise PoleImportError(f"Could not open Excel file: {error}") from error
    try:
        return [tuple(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()


def _detect_header_index(rows: list[tuple[object, ...]]) -> int:
    candidates = rows[:HEADER_SCAN_LIMIT]
    scores = [sum(_recognized_field(value) is not None for value in row) for row in candidates]
    best_index = max(range(len(candidates)), key=scores.__getitem__)
    if scores[best_index] == 0:
        return next(
            (index for index, row in enumerate(candidates) if any(value not in (None, "") for value in row)),
            0,
        )
    return best_index


def _recognized_field(value: object) -> str | None:
    normalized = _normalize_header(value)
    return next((field for field, aliases in HEADER_ALIASES.items() if normalized in aliases), None)


def _normalize_header(value: object) -> str:
    return re.sub(r"[\W_]+", "", str(value or "").strip().casefold(), flags=re.UNICODE)


def _unique_headers(row: tuple[object, ...]) -> tuple[str, ...]:
    headers: list[str] = []
    for index, value in enumerate(row, start=1):
        base = str(value).strip() if value not in (None, "") else f"Column {index}"
        header = base
        suffix = 2
        while header in headers:
            header = f"{base} ({suffix})"
            suffix += 1
        headers.append(header)
    return tuple(headers)


def _row_to_pole(
    row: tuple[object, ...],
    row_number: int,
    mapping: dict[str, str | None],
    header_indexes: dict[str, int],
) -> Pole:
    def value(field: str) -> object:
        column = mapping.get(field)
        return row[header_indexes[column]] if column else None

    try:
        return Pole(
            number=str(value("number") or "").strip(),
            latitude=float(value("latitude")),
            longitude=float(value("longitude")),
            detail=str(value("detail") or "").strip(),
            side=PoleSide.from_text(value("side")),
        )
    except (TypeError, ValueError) as error:
        raise PoleImportError(f"Row {row_number}: {error}") from error

