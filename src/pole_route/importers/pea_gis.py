"""Discover and import profile-driven data from PEA GIS workbooks."""

from __future__ import annotations

import math
import re
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pole_route.domain.pea_gis import PEAPoleRecord


class PEAGISImportError(ValueError):
    """A PEA GIS workbook cannot be inspected or imported safely."""


@dataclass(frozen=True, slots=True)
class PEASheetDiscovery:
    name: str
    profile: str | None

    @property
    def supported(self) -> bool:
        return self.profile is not None


@dataclass(frozen=True, slots=True)
class PEAWorkbookDiscovery:
    sheets: tuple[PEASheetDiscovery, ...]

    @property
    def sheet_names(self) -> tuple[str, ...]:
        return tuple(sheet.name for sheet in self.sheets)

    @property
    def supported_sheets(self) -> tuple[PEASheetDiscovery, ...]:
        return tuple(sheet for sheet in self.sheets if sheet.supported)

    @property
    def unsupported_ds_sheets(self) -> tuple[str, ...]:
        return tuple(
            sheet.name
            for sheet in self.sheets
            if sheet.name.casefold().startswith("ds_") and not sheet.supported
        )


DS_POLE_PROFILE = "DS_Pole"
DS_TRANSFORMER_PROFILE = "DS_Transformer"
DS_SWITCH_PROFILE = "DS_Switch"
HEADER_SCAN_LIMIT = 30
THAI_DIGITS = str.maketrans("๐๑๒๓๔๕๖๗๘๙", "0123456789")

DS_POLE_ALIASES = {
    "source_id": {
        "poleid", "pole_id", "poleno", "pole_no", "peano", "pea_no",
        "หมายเลขเสา", "เลขเสา", "รหัสเสา", "รหัสทรัพย์สิน",
    },
    "latitude": {"latitude", "lat", "ละติจูด"},
    "longitude": {"longitude", "long", "lon", "lng", "ลองจิจูด"},
    "height": {"height", "poleheight", "pole_height", "heightm", "ความสูง", "ความสูงเสา"},
    "voltage": {"voltage", "volt", "voltagelevel", "voltage_level", "แรงดัน", "ระดับแรงดัน"},
}


def discover_pea_workbook(path: str | Path) -> PEAWorkbookDiscovery:
    source = _validated_source(path)
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as error:
        raise PEAGISImportError(f"Could not open PEA GIS workbook: {error}") from error
    try:
        profiles = {
            "ds_pole": DS_POLE_PROFILE,
            "ds_transformer": DS_TRANSFORMER_PROFILE,
            "ds_switch": DS_SWITCH_PROFILE,
        }
        sheets = tuple(
            PEASheetDiscovery(name, profiles.get(name.casefold()))
            for name in workbook.sheetnames
        )
    finally:
        workbook.close()
    return PEAWorkbookDiscovery(sheets)


def import_ds_poles(path: str | Path) -> list[PEAPoleRecord]:
    source = _validated_source(path)
    discovery = discover_pea_workbook(source)
    sheet_name = next(
        (sheet.name for sheet in discovery.supported_sheets if sheet.profile == DS_POLE_PROFILE),
        None,
    )
    if sheet_name is None:
        visible = ", ".join(discovery.sheet_names) or "(no sheets)"
        raise PEAGISImportError(
            f"DS_Pole worksheet was not found. Workbook sheets: {visible}"
        )

    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as error:
        raise PEAGISImportError(f"Could not open PEA GIS workbook: {error}") from error
    try:
        worksheet = workbook[sheet_name]
        rows = [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()
    if not rows:
        raise PEAGISImportError("DS_Pole worksheet is empty")

    header_index, headers, mapping = _resolve_ds_pole_headers(rows)
    records: list[PEAPoleRecord] = []
    for offset, row in enumerate(rows[header_index + 1 :], start=1):
        if not any(value not in (None, "") for value in row):
            continue
        records.append(
            _parse_ds_pole_row(
                headers,
                row,
                mapping,
                source_sheet=sheet_name,
                source_row=header_index + offset + 1,
            )
        )
    if not records:
        raise PEAGISImportError("DS_Pole contains headers but no data rows")
    return records


def parse_height_metres(value: Any) -> float | None:
    if value in (None, "") or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        numeric = float(value)
    else:
        text = str(value).strip().translate(THAI_DIGITS).casefold()
        match = re.fullmatch(r"([+-]?\d+(?:\.\d+)?)\s*(?:m|metre|meter|เมตร)?", text)
        if not match:
            return None
        numeric = float(match.group(1))
    return numeric if math.isfinite(numeric) and numeric > 0 else None


def parse_voltage_kv(value: Any) -> tuple[float | None, float | None]:
    """Return a normalized voltage span in kV, preserving raw data elsewhere."""
    if value in (None, "") or isinstance(value, bool):
        return None, None
    text = str(value).strip().translate(THAI_DIGITS).casefold()
    unit_matches = list(re.finditer(r"\b(kv|v)\b", text, re.IGNORECASE))
    if not unit_matches:
        return None, None
    numbers = [float(item) for item in re.findall(r"\d+(?:\.\d+)?", text)]
    if not numbers or len(numbers) > 2:
        return None, None
    # PEA values normally put one unit after a slash/range. If individual units
    # occur, apply the nearest following unit to each numeric value.
    values: list[float] = []
    number_matches = list(re.finditer(r"\d+(?:\.\d+)?", text))
    for index, number_match in enumerate(number_matches):
        following = next((match for match in unit_matches if match.start() >= number_match.end()), None)
        unit = following.group(1).casefold() if following else unit_matches[-1].group(1).casefold()
        numeric = numbers[index] / 1000.0 if unit == "v" else numbers[index]
        values.append(numeric)
    return min(values), max(values)


def default_pole_inclusion(height_metres: float | None, voltage_max_kv: float | None) -> bool:
    return bool(
        (height_metres is not None and height_metres >= 12.0)
        or (voltage_max_kv is not None and voltage_max_kv >= 22.0)
    )


def _validated_source(path: str | Path) -> Path:
    source = Path(path)
    if source.suffix.casefold() != ".xlsx":
        raise PEAGISImportError("Choose a .xlsx PEA GIS workbook")
    if not source.is_file():
        raise PEAGISImportError(f"File not found: {source}")
    return source


def _resolve_ds_pole_headers(
    rows: list[tuple[Any, ...]],
) -> tuple[int, tuple[str, ...], dict[str, int]]:
    normalized_aliases = {
        field: {_normalize_header(alias) for alias in aliases}
        for field, aliases in DS_POLE_ALIASES.items()
    }
    best: tuple[int, tuple[str, ...], dict[str, int]] | None = None
    for row_index, row in enumerate(rows[:HEADER_SCAN_LIMIT]):
        headers = _unique_headers(row)
        mapping: dict[str, int] = {}
        for column, header in enumerate(headers):
            normalized = _normalize_header(header)
            for field, aliases in normalized_aliases.items():
                if normalized in aliases and field not in mapping:
                    mapping[field] = column
        if best is None or len(mapping) > len(best[2]):
            best = row_index, headers, mapping
    assert best is not None
    missing = [field for field in DS_POLE_ALIASES if field not in best[2]]
    if missing:
        labels = {
            "source_id": "Pole ID", "latitude": "Latitude", "longitude": "Longitude",
            "height": "Height", "voltage": "Voltage",
        }
        raise PEAGISImportError(
            "DS_Pole required columns could not be resolved: "
            + ", ".join(labels[field] for field in missing)
        )
    return best


def _parse_ds_pole_row(
    headers: tuple[str, ...],
    row: tuple[Any, ...],
    mapping: dict[str, int],
    *,
    source_sheet: str,
    source_row: int,
) -> PEAPoleRecord:
    padded = tuple(row[index] if index < len(row) else None for index in range(len(headers)))
    raw_attributes = {headers[index]: _json_safe(value) for index, value in enumerate(padded)}
    source_id = str(padded[mapping["source_id"]] or "").strip()
    if not source_id:
        raise PEAGISImportError(f"{source_sheet} row {source_row}: Pole ID is required")
    latitude = _coordinate(padded[mapping["latitude"]], "Latitude", source_sheet, source_row)
    longitude = _coordinate(padded[mapping["longitude"]], "Longitude", source_sheet, source_row)
    raw_height = _json_safe(padded[mapping["height"]])
    raw_voltage = _json_safe(padded[mapping["voltage"]])
    height = parse_height_metres(raw_height)
    voltage_min, voltage_max = parse_voltage_kv(raw_voltage)
    warnings: list[str] = []
    if height is None:
        warnings.append("Pole height could not be normalized")
    if voltage_min is None or voltage_max is None:
        warnings.append("Voltage could not be normalized")
    return PEAPoleRecord(
        source_id=source_id,
        source_sheet=source_sheet,
        source_row=source_row,
        latitude=latitude,
        longitude=longitude,
        raw_height=raw_height,
        height_metres=height,
        raw_voltage=raw_voltage,
        voltage_min_kv=voltage_min,
        voltage_max_kv=voltage_max,
        raw_attributes=raw_attributes,
        included_by_default=default_pole_inclusion(height, voltage_max),
        qc_warnings=tuple(warnings),
    )


def _coordinate(value: Any, label: str, sheet: str, row: int) -> float:
    try:
        numeric = float(str(value).strip().translate(THAI_DIGITS).rstrip("°"))
    except (TypeError, ValueError) as error:
        raise PEAGISImportError(f"{sheet} row {row}: {label} is invalid") from error
    limit = 90 if label == "Latitude" else 180
    if not math.isfinite(numeric) or not -limit <= numeric <= limit:
        raise PEAGISImportError(f"{sheet} row {row}: {label} is out of range")
    return numeric


def _normalize_header(value: Any) -> str:
    return re.sub(r"[\W_]+", "", str(value or "").strip().casefold(), flags=re.UNICODE)


def _unique_headers(row: Iterable[Any]) -> tuple[str, ...]:
    result: list[str] = []
    for index, value in enumerate(row, start=1):
        base = str(value).strip() if value not in (None, "") else f"Column {index}"
        candidate = base
        suffix = 2
        while candidate in result:
            candidate = f"{base} ({suffix})"
            suffix += 1
        result.append(candidate)
    return tuple(result)


def _json_safe(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)
