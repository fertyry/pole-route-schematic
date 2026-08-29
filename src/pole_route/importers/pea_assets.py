"""Profile-driven import of coordinate-bearing PEA GIS assets."""

from __future__ import annotations

import hashlib
import json
import math
from collections.abc import Mapping
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from pole_route.domain.pea_asset import PEAAsset, PEAAssetType
from pole_route.importers.pea_gis import (
    HEADER_SCAN_LIMIT,
    THAI_DIGITS,
    PEAGISImportError,
    _json_safe,
    _normalize_header,
    _unique_headers,
    _validated_source,
    parse_voltage_kv,
)


@dataclass(frozen=True, slots=True)
class PEAAssetProfile:
    profile_id: str
    sheet_name: str
    asset_type: PEAAssetType
    aliases: Mapping[str, frozenset[str]]


COMMON_ALIASES = {
    "latitude": frozenset({"latitude", "lat", "ละติจูด"}),
    "longitude": frozenset({"longitude", "long", "lon", "lng", "ลองจิจูด"}),
    "name": frozenset({"name", "label", "ชื่อ", "ชื่ออุปกรณ์"}),
    "voltage": frozenset({"voltage", "volt", "voltagelevel", "แรงดัน", "ระดับแรงดัน"}),
    "phase": frozenset({"phase", "เฟส", "เฟสที่ติดตั้ง"}),
    "status": frozenset({"status", "สถานะ", "สถานะปัจจุบัน", "สถานะการก่อสร้าง"}),
    "feeder": frozenset({
        "feeder", "feederid", "circuit", "วงจร", "ฟีดเดอร์", "รหัสสายป้อนที่ 1",
    }),
}


DS_TRANSFORMER_PROFILE = PEAAssetProfile(
    "DS_Transformer",
    "DS_Transformer",
    PEAAssetType.TRANSFORMER,
    {
        **COMMON_ALIASES,
        "source_id": frozenset({
            "transformerid", "transformer_id", "equipmentid", "equipment_id",
            "assetid", "asset_id", "รหัสหม้อแปลง", "หมายเลขหม้อแปลง", "รหัสทรัพย์สิน",
            "PEANO หม้อแปลง", "PEANO (ADS)",
        }),
        "rating": frozenset({
            "rating", "capacity", "kva", "ขนาด", "พิกัด", "กำลัง",
            "ค่าพิกัด kVA หม้อแปลง",
        }),
        "subtype": frozenset({
            "type", "transformertype", "equipmenttype", "ชนิด", "ประเภท",
            "ประเภทการติดตั้ง",
        }),
    },
)

DS_SWITCH_PROFILE = PEAAssetProfile(
    "DS_Switch",
    "DS_Switch",
    PEAAssetType.SWITCH,
    {
        **COMMON_ALIASES,
        "source_id": frozenset({
            "switchid", "switch_id", "equipmentid", "equipment_id", "assetid",
            "asset_id", "รหัสสวิตช์", "หมายเลขสวิตช์", "รหัสอุปกรณ์", "รหัสทรัพย์สิน",
        }),
        "rating": frozenset({"rating", "capacity", "ampere", "amp", "พิกัด", "กระแส"}),
        "subtype": frozenset({
            "type", "switchtype", "equipmenttype", "ชนิด", "ประเภท",
            "ประเภทย่อยของสวิตช์",
        }),
    },
)

ASSET_PROFILES = (DS_TRANSFORMER_PROFILE, DS_SWITCH_PROFILE)
ASSET_PROFILE_BY_SHEET = {item.sheet_name.casefold(): item for item in ASSET_PROFILES}


def import_pea_assets(
    path: str | Path,
    profiles: tuple[PEAAssetProfile, ...] | None = None,
) -> list[PEAAsset]:
    """Import every row from selected supported sheets, retaining QC failures."""
    source = _validated_source(path)
    selected = profiles or ASSET_PROFILES
    try:
        workbook = load_workbook(source, read_only=True, data_only=True)
    except Exception as error:
        raise PEAGISImportError(f"Could not open PEA GIS workbook: {error}") from error
    assets: list[PEAAsset] = []
    try:
        names = {name.casefold(): name for name in workbook.sheetnames}
        for profile in selected:
            actual_name = names.get(profile.sheet_name.casefold())
            if actual_name is None:
                continue
            rows = [tuple(row) for row in workbook[actual_name].iter_rows(values_only=True)]
            if not rows:
                continue
            header_index, headers, mapping = _resolve_headers(rows, profile)
            for offset, row in enumerate(rows[header_index + 1 :], start=1):
                if any(value not in (None, "") for value in row):
                    assets.append(_parse_row(
                        headers, row, mapping, profile,
                        source_sheet=actual_name,
                        source_row=header_index + offset + 1,
                    ))
    finally:
        workbook.close()
    return assets


def _resolve_headers(rows: list[tuple[Any, ...]], profile: PEAAssetProfile):
    normalized = {
        field: {_normalize_header(alias) for alias in aliases}
        for field, aliases in profile.aliases.items()
    }
    best = None
    for row_index, row in enumerate(rows[:HEADER_SCAN_LIMIT]):
        headers = _unique_headers(row)
        mapping: dict[str, int] = {}
        for column, header in enumerate(headers):
            key = _normalize_header(header)
            for field, aliases in normalized.items():
                if key in aliases and field not in mapping:
                    mapping[field] = column
        if best is None or len(mapping) > len(best[2]):
            best = (row_index, headers, mapping)
    assert best is not None
    missing = [field for field in ("latitude", "longitude") if field not in best[2]]
    if missing:
        raise PEAGISImportError(
            f"{profile.sheet_name} required columns could not be resolved: "
            + ", ".join(field.title() for field in missing)
        )
    return best


def _text(value: Any) -> str | None:
    if value in (None, ""):
        return None
    result = str(value).strip()
    return result or None


def _optional_coordinate(value: Any, label: str) -> tuple[float | None, str | None]:
    try:
        numeric = float(str(value).strip().translate(THAI_DIGITS).rstrip("°"))
    except (TypeError, ValueError):
        return None, f"{label} is missing or invalid"
    limit = 90 if label == "Latitude" else 180
    if not math.isfinite(numeric) or not -limit <= numeric <= limit:
        return None, f"{label} is out of range"
    return numeric, None


def _parse_row(headers, row, mapping, profile, *, source_sheet, source_row):
    padded = tuple(row[index] if index < len(row) else None for index in range(len(headers)))
    raw = {headers[index]: _json_safe(value) for index, value in enumerate(padded)}
    value = lambda field: padded[mapping[field]] if field in mapping else None
    source_id = _text(value("source_id")) or ""
    latitude, lat_warning = _optional_coordinate(value("latitude"), "Latitude")
    longitude, lon_warning = _optional_coordinate(value("longitude"), "Longitude")
    warnings = [item for item in (lat_warning, lon_warning) if item]
    if not source_id:
        warnings.append("Source asset ID is missing; content fingerprint identity is used")
    fingerprint_data = {
        _normalize_header(key): raw[key]
        for key in sorted(raw, key=_normalize_header)
        if raw[key] not in (None, "")
    }
    fingerprint = hashlib.sha256(
        json.dumps(fingerprint_data, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()[:20]
    identity = source_id.casefold() if source_id else f"fingerprint:{fingerprint}"
    raw_voltage = _json_safe(value("voltage"))
    voltage_min, voltage_max = parse_voltage_kv(raw_voltage)
    return PEAAsset(
        stable_id=f"{profile.profile_id}:{identity}",
        source_sheet=source_sheet,
        source_row=source_row,
        asset_type=profile.asset_type,
        source_asset_id=source_id,
        latitude=latitude,
        longitude=longitude,
        raw_attributes=raw,
        name=_text(value("name")),
        raw_voltage=raw_voltage,
        voltage_min_kv=voltage_min,
        voltage_max_kv=voltage_max,
        rating=_text(value("rating")),
        phase=_text(value("phase")),
        equipment_subtype=_text(value("subtype")),
        status=_text(value("status")),
        feeder_reference=_text(value("feeder")),
        qc_warnings=tuple(warnings),
    )
