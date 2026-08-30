"""Shared Qt-free worksheet discovery and evidence-based header detection."""

from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
SUPPORTED_SUFFIXES = {".csv", *EXCEL_SUFFIXES}
HEADER_SCAN_LIMIT = 30


@dataclass(frozen=True, slots=True)
class HeaderDetection:
    row_index: int
    confidence: str
    mapping: dict[str, int]
    score: int

    @property
    def row_number(self) -> int:
        return self.row_index + 1


def list_worksheets(path: str | Path) -> tuple[str, ...]:
    source = Path(path)
    if source.suffix.casefold() == ".csv":
        return ()
    if source.suffix.casefold() not in EXCEL_SUFFIXES:
        raise ValueError("Choose a .csv, .xlsx, or .xlsm file")
    workbook = load_workbook(source, read_only=True, data_only=True, keep_vba=False)
    try:
        return tuple(workbook.sheetnames)
    finally:
        workbook.close()


def read_rows(path: str | Path, sheet_name: str | None = None) -> list[tuple[object, ...]]:
    source = Path(path)
    suffix = source.suffix.casefold()
    if suffix == ".csv":
        try:
            with source.open("r", encoding="utf-8-sig", newline="") as handle:
                return [tuple(row) for row in csv.reader(handle)]
        except UnicodeDecodeError as error:
            raise ValueError("CSV must be saved with UTF-8 encoding") from error
    if suffix not in EXCEL_SUFFIXES:
        raise ValueError("Choose a .csv, .xlsx, or .xlsm file")
    workbook = load_workbook(source, read_only=True, data_only=True, keep_vba=False)
    try:
        if sheet_name is not None and sheet_name not in workbook.sheetnames:
            raise ValueError(f"Worksheet not found: {sheet_name}")
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        return [tuple(row) for row in worksheet.iter_rows(values_only=True)]
    finally:
        workbook.close()


def detect_header(
    rows: list[tuple[object, ...]],
    aliases: dict[str, set[str]],
    required_fields: tuple[str, ...],
) -> HeaderDetection:
    if not rows:
        raise ValueError("The selected worksheet is empty")
    normalized_aliases = {
        field: {normalize_header(alias) for alias in values}
        for field, values in aliases.items()
    }
    candidates: list[tuple[int, int, dict[str, int], int]] = []
    for row_index, row in enumerate(rows[:HEADER_SCAN_LIMIT]):
        mapping: dict[str, int] = {}
        non_empty = 0
        for column, value in enumerate(row):
            if value not in (None, ""):
                non_empty += 1
            normalized = normalize_header(value)
            for field, values in normalized_aliases.items():
                if normalized in values and field not in mapping:
                    mapping[field] = column
        required_count = sum(field in mapping for field in required_fields)
        coordinate_pair = int("latitude" in mapping and "longitude" in mapping)
        score = required_count * 10 + coordinate_pair * 8 + len(mapping) * 2 + min(non_empty, 8)
        candidates.append((score, -row_index, mapping, row_index))
    score, _, mapping, row_index = max(candidates)
    required_count = sum(field in mapping for field in required_fields)
    if required_count == len(required_fields):
        confidence = "High"
    elif "latitude" in mapping and "longitude" in mapping and required_count >= 2:
        confidence = "Medium"
    else:
        confidence = "Low"
    return HeaderDetection(row_index, confidence, mapping, score)


def unique_headers(row: tuple[object, ...]) -> tuple[str, ...]:
    headers: list[str] = []
    for index, value in enumerate(row, start=1):
        base = str(value).strip() if value not in (None, "") else f"Column {index}"
        header, suffix = base, 2
        while header in headers:
            header = f"{base} ({suffix})"
            suffix += 1
        headers.append(header)
    return tuple(headers)


def normalize_header(value: object) -> str:
    return re.sub(r"[\W_]+", "", str(value or "").strip().casefold(), flags=re.UNICODE)
